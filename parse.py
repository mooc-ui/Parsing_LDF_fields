
#!/usr/bin/env python3
"""
LIN LDF Signal Mapping Visualizer
解析 LDF 文件并生成类似 CANoe/Vector 风格的 Signal Mapping 图
支持: CLI 直接运行 (python parse.py) 或作为模块被 gui.py 调用
"""

import ldfparser
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle, Polygon, PathPatch, Circle
from matplotlib.path import Path
from matplotlib.colors import LinearSegmentedColormap
import os
import math
import textwrap

# ==================== 中文字体配置 ====================
plt.rcParams["font.family"] = "sans-serif"
plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "Noto Sans SC", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False

# ==================== 配置 ====================
LDF_PATH = "./LDF_EESystem25_23KW51_V36_M_LIN_4_V297.ldf"
OUTPUT_DIR = "./output"
DEFAULT_FRAMES = [
    "ST_FAN_1_LIN",
    "CTR_FAN_LIN",
    "StatusThreeWayValveBatt48LIN",
    "ControlThreeWayValveBatt48LIN",
]


# ==================== 工具函数 ====================
def get_signals(frame):
    """获取帧内信号列表（按起始位排序）"""
    signals = []
    for start_bit, sig in sorted(frame.signal_map, key=lambda x: x[0]):
        signals.append({
            "name": sig.name,
            "start": start_bit,
            "size": sig.width,
            "end": start_bit + sig.width - 1
        })
    return signals


def draw_signal_label(ax, sig, y, base_fs=6.5, min_fs=3.0):
    """绘制信号完整标识符（不简写），窄字段或长名称时自动旋转/缩字体"""
    name = sig["name"]
    rot = 90 if sig["size"] < 8 else 0
    width_lim = base_fs * sig["size"] / 10.0 if rot == 0 else base_fs * sig["size"] / max(len(name), 4)
    len_lim = base_fs * 10.0 / max(len(name), 10) + 0.5
    fs = max(min_fs, min(base_fs, width_lim, len_lim))
    ax.text(sig["start"] + sig["size"] / 2, y, name, ha="center", va="center",
            rotation=rot, fontsize=fs, fontweight="normal", color="#111", zorder=5)


def plot_single_frame(ldf, frame_name, output_dir=OUTPUT_DIR, save_png=True, save_svg=True, dpi=300,
                      annotations=None):
    """绘制单帧 Signal Mapping 图，返回生成的文件路径列表

    annotations: dict {信号名: 注释文本}，注释会以引线+文本框绘制在信号块下方
    """
    frame = ldf.get_unconditional_frame(frame_name)
    signals = get_signals(frame)
    total_bits = frame.length * 8

    # 收集该帧的注释（按信号顺序，与信号块保持对应）
    ann_items = []
    if annotations:
        for sig in signals:
            text = (annotations.get(sig["name"]) or "").strip()
            if text:
                ann_items.append((sig, text))

    # 注释区高度：每个注释一行(0.32)，多行注释按行数扩展
    ann_rows = 0
    for _, text in ann_items:
        n_lines = len(textwrap.wrap(text, 28)) if any('\u4e00' <= ch <= '\u9fff' for ch in text) \
            else len(textwrap.wrap(text, 50))
        ann_rows += n_lines
    ann_height = 0.32 * max(ann_rows, len(ann_items))

    # 颜色渐变（黄绿 → 青绿 → 深蓝）
    colors = ["#f0e68c", "#c5e1a5", "#81c784", "#4db6ac", "#26a69a",
              "#00897b", "#00695c", "#004d40", "#1a237e", "#0d47a1", "#ffcc80"]
    cmap = LinearSegmentedColormap.from_list("custom", colors, N=len(signals))

    fig_h = 5 + ann_height
    fig, ax = plt.subplots(figsize=(max(8, frame.length * 2.25), fig_h))
    ax.set_facecolor("#fafafa")

    # 垂直网格线
    for i in range(total_bits + 1):
        lw = 1.2 if i % 8 == 0 else 0.4
        color = "#333" if i % 8 == 0 else "#ccc"
        ax.axvline(i, color=color, linewidth=lw, ymin=0.15, ymax=0.85, zorder=1)

    # 绘制信号块
    for i, sig in enumerate(signals):
        color = cmap(i / max(len(signals) - 1, 1))
        rect = Rectangle((sig["start"], 0.35), sig["size"], 0.3,
                         facecolor=color, edgecolor="#222", linewidth=0.6, alpha=0.9, zorder=2)
        ax.add_patch(rect)

        # 信号完整标识符（不简写）
        draw_signal_label(ax, sig, 0.72)

    # 中间高亮斜线（仅 ST_FAN_1_LIN 原图风格，对应 AVL_U 区域）
    if frame.name == "ST_FAN_1_LIN":
        hatch_rect = Rectangle((24, 0.35), 8, 0.3, facecolor="none", edgecolor="#555",
                               linewidth=0.8, hatch="///", alpha=0.6, zorder=3)
        ax.add_patch(hatch_rect)

    # 底部箭头
    for i in range(0, total_bits, 2):
        ax.annotate("", xy=(i + 1.5, 0.28), xytext=(i + 0.5, 0.28),
                    arrowprops=dict(arrowstyle="->", color="black", lw=0.8), zorder=4)

    # 顶部字节号
    for i in range(frame.length):
        ax.text(i * 8 + 4, 0.95, str(i), ha="center", va="center", fontsize=11, fontweight="bold",
                bbox=dict(boxstyle="round,pad=0.2", facecolor="white", edgecolor="none", alpha=0.8))

    # 底部 bit 编号
    for i in range(0, total_bits, 8):
        ax.text(i, 0.08, str(i), ha="center", va="top", fontsize=8, color="#333")
    for i in range(7, total_bits, 8):
        ax.text(i, 0.08, str(i), ha="center", va="top", fontsize=7, color="#666")

    ax.set_title(f"Signal Mapping — {frame.name}  (Frame ID: {frame.frame_id}, Length: {frame.length} bytes)",
                 fontsize=13, fontweight="bold", pad=10)
    ax.set_xlim(-0.5, total_bits + 0.5)
    ax.set_ylim(-ann_height - 0.12 if ann_items else 0, 1.1)
    ax.set_yticks([])
    ax.set_xticks([])
    for spine in ax.spines.values():
        spine.set_visible(False)

    # ===== 绘制注释（引线 + 文本框，位于信号块下方） =====
    y_cursor = -0.08
    for sig, text in ann_items:
        cx = sig["start"] + sig["size"] / 2
        # 引线：从信号块下边缘到注释框上边缘
        ax.plot([cx, cx], [0.35, y_cursor + 0.20], color="#c0392b", lw=0.9, zorder=4)

        # 注释文本框
        wrap_w = 28 if any('\u4e00' <= ch <= '\u9fff' for ch in text) else 50
        wrapped = "\n".join(textwrap.wrap(text, wrap_w)) or text
        ax.text(cx, y_cursor + 0.16, wrapped, ha="center", va="center", fontsize=8,
                color="#222", zorder=5,
                bbox=dict(boxstyle="round,pad=0.35", facecolor="#fff8e1",
                          edgecolor="#c0392b", lw=0.8))
        n_lines = len(wrapped.split("\n"))
        y_cursor -= 0.22 * n_lines + 0.10

    plt.tight_layout()

    saved = []
    if save_png:
        png_path = f"{output_dir}/{frame_name}_signal_mapping.png"
        plt.savefig(png_path, dpi=dpi, bbox_inches="tight", facecolor="white")
        saved.append(png_path)
    if save_svg:
        svg_path = f"{output_dir}/{frame_name}_signal_mapping.svg"
        plt.savefig(svg_path, format="svg", bbox_inches="tight", facecolor="white")
        saved.append(svg_path)
    plt.close()
    return saved


# ==================== 八边形布局 ====================
def _octagon_geometry():
    """正八边形几何：顶点(8个)与字节边定义(byte b -> (起点顶点, 终点顶点))
    顶点 k 角度 = 67.5° + 45°*k；byte0=顶边(左→右)，顺时针排列"""
    R = 1.0
    verts = []
    for k in range(8):
        ang = math.radians(67.5 + 45 * k)
        verts.append((R * math.cos(ang), R * math.sin(ang)))
    byte_edges = [(1, 0), (0, 7), (7, 6), (6, 5), (5, 4), (4, 3), (3, 2), (2, 1)]
    return verts, byte_edges


def _octagon_edge_point(verts, byte_edges, b, t):
    """字节边 b 上参数 t (0~1) 处的坐标（t=0 起点顶点, t=1 终点顶点）"""
    ks, ke = byte_edges[b]
    x0, y0 = verts[ks]
    x1, y1 = verts[ke]
    return (x0 + (x1 - x0) * t, y0 + (y1 - y0) * t)


def _octagon_outward(verts, byte_edges, b):
    """字节边 b 的外法向（单位向量），正八边形边中点径向即外法向"""
    mx, my = _octagon_edge_point(verts, byte_edges, b, 0.5)
    L = math.hypot(mx, my)
    return (mx / L, my / L)


def _octagon_block_quad(verts, byte_edges, b, t0, t1, off, width):
    """信号块四边形：字节边 b 上 [t0,t1] 段，沿外法向偏移 off~off+width"""
    p0 = _octagon_edge_point(verts, byte_edges, b, t0)
    p1 = _octagon_edge_point(verts, byte_edges, b, t1)
    nx, ny = _octagon_outward(verts, byte_edges, b)
    return [(p0[0] + nx * off, p0[1] + ny * off),
            (p1[0] + nx * off, p1[1] + ny * off),
            (p1[0] + nx * (off + width), p1[1] + ny * (off + width)),
            (p0[0] + nx * (off + width), p0[1] + ny * (off + width))]


def _split_byte_segments(sig):
    """把信号拆成按字节分段的 [(byte, j0, j1)]，j0/j1 为该字节内 bit 序号(含两端)"""
    segs = []
    s, e = sig["start"], sig["end"]
    for b in range(s // 8, e // 8 + 1):
        j0 = s if b == s // 8 else b * 8
        j1 = e if b == e // 8 else b * 8 + 7
        segs.append((b, j0 - b * 8, j1 - b * 8))
    return segs


def _measure_text_plain_r(ax, text, fs, renderer):
    """测量文本（无边框）宽度，单位：axes 归一化坐标（0~1）。

    使用已获取的 renderer 直接计算窗口尺寸，不触发整图重绘，速度快。
    """
    t = ax.text(0, 0, text, ha="left", va="center", fontsize=fs)
    bb = t.get_window_extent(renderer=renderer)
    inv = ax.transData.inverted()
    x0, _ = inv.transform((bb.x0, bb.y0))
    x1, _ = inv.transform((bb.x1, bb.y1))
    t.remove()
    return x1 - x0


def _layout_octagon_annotation_list(ax, ann_items, sig_colors, fs_start=10.0, y_top=0.94):
    """注释列表式布局：右侧区域从上到下一条条顺序排列，放不下自动分列。

    每条注释 = 色块 + 加粗信号名(独立一行) + 注释文本(缩进换行)，
    无论名称多长都不会横向越界/与相邻条目重叠。
    ax: 归一化坐标(0~1)的注释区 axes
    返回 (entries, fs, line_h)：
      - entries: [(sig, name, name_w, text_lines, x, y, h, color)]
      - line_h: 单行行距（归一化高度）
    """
    renderer = ax.figure.canvas.get_renderer()
    # 单行高度（归一化）
    t = ax.text(0, 0, "中", fontsize=fs_start, ha="left", va="center")
    bb = t.get_window_extent(renderer=renderer)
    inv = ax.transData.inverted()
    y0 = inv.transform((bb.x0, bb.y0))[1]
    y1 = inv.transform((bb.x0, bb.y1))[1]
    t.remove()
    base_line = y1 - y0

    def wrap_to_avail(text, avail, fs):
        lines = []
        cur = ""
        for ch in text:
            test = cur + ch
            if cur and _measure_text_plain_r(ax, test, fs, renderer) > avail:
                lines.append(cur)
                cur = ch
            else:
                cur = test
        if cur:
            lines.append(cur)
        return lines

    best = None   # (max_col_h, fs, n_cols, col_xs, col_items)
    for fs in (fs_start, fs_start - 1.0, fs_start - 2.0, fs_start - 3.0):
        line_h = base_line * (fs / fs_start) * 1.45
        gap = base_line * (fs / fs_start) * 0.6
        for n_cols in (1, 2, 3, 4):
            col_w = 0.94 if n_cols == 1 else (0.455 if n_cols == 2
                                              else (0.30 if n_cols == 3 else 0.225))
            col_xs = [0.02, 0.52, 0.68, 0.76][:n_cols]
            indent = 0.035             # 文本相对名称缩进
            avail = col_w - indent - 0.02
            if avail < base_line * 0.5:   # 列太窄放不下任何字 → 换更大列数/更小字号
                continue
            rows = []
            for sig, text in ann_items:
                # 名称行 = 信号名 + bit 范围（从哪一bit到哪一bit）
                name = f"{sig['name']}  [bit {sig['start']}-{sig['end']}]"
                name_w = _measure_text_plain_r(ax, name, fs, renderer)
                # 名称独立一行：若名称超宽则按列宽换行（保证不横向越界）
                if name_w > col_w - 0.02:
                    name_lines = wrap_to_avail(name, col_w - 0.02, fs)
                    name_w = col_w - 0.02
                    disp_name = name_lines[0]
                else:
                    disp_name = name
                lines = wrap_to_avail(text, avail, fs)
                h = (1 + len(lines)) * line_h + gap
                rows.append((sig, disp_name, name_w, lines, h))
            # 贪心分配列：依次放入当前最矮列
            col_h = [0.0] * n_cols
            col_items = [[] for _ in range(n_cols)]
            for row in rows:
                ci = min(range(n_cols), key=lambda i: col_h[i])
                col_items[ci].append(row)
                col_h[ci] += row[4]
            mh = max(col_h)
            if mh <= 0.94:
                entries = []
                for ci in range(n_cols):
                    y = y_top
                    for sig, name, name_w, lines, h in col_items[ci]:
                        entries.append((sig, name, name_w, lines,
                                        col_xs[ci], y, h,
                                        sig_colors.get(sig["name"], "#888")))
                        y -= h
                return entries, fs, line_h
            if best is None or mh < best[0]:
                best = (mh, fs, n_cols, col_xs, col_items)
    # 极端情况兜底：取高度最小方案，保证文本换行完整（不重叠）
    _, fs, n_cols, col_xs, col_items = best
    line_h = base_line * (fs / fs_start) * 1.45
    entries = []
    for ci in range(n_cols):
        y = y_top
        for sig, name, name_w, lines, h in col_items[ci]:
            entries.append((sig, name, name_w, lines,
                            col_xs[ci], y, h,
                            sig_colors.get(sig["name"], "#888")))
            y -= h
    return entries, fs, line_h


def _draw_octagon_annotation_list(ax, entries, fs, line_h):
    """绘制注释列表条目：色块 + 加粗信号名(首行) + 注释文本(从第二行缩进)"""
    sw = line_h * 0.55
    for sig, name, name_w, lines, x, y, h, color in entries:
        # 色块（位于条目顶部行，垂直居中于名称行）
        ax.add_patch(Rectangle((x, y + line_h * 0.35 - sw), sw, sw,
                               transform=ax.transAxes,
                               facecolor=color, edgecolor="#333", lw=0.4, zorder=6))
        tx = x + sw + 0.02
        # 信号名（加粗，首行）
        ax.text(tx, y, name, ha="left", va="top", fontsize=fs,
                fontweight="bold", color="#111", transform=ax.transAxes, zorder=7)
        # 注释文本：从第二行起缩进（不与名称重叠）
        yy = y + line_h
        for ln in lines:
            ax.text(tx + 0.02, yy, ln, ha="left", va="top", fontsize=fs,
                    color="#333", transform=ax.transAxes, zorder=7)
            yy += line_h


def _measure_text(ax, text, fs):
    """测量文本（含 bbox 边框）在数据坐标系中的实际渲染尺寸 (w, h)。

    依赖当前 axes 的变换（xlim/ylim/tight_layout 已就位），用于防重叠布局。
    """
    fig = ax.figure
    inv = ax.transData.inverted()
    t = ax.text(0, 0, text, ha="center", va="center", fontsize=fs,
                bbox=dict(boxstyle="round,pad=0.3", facecolor="#fff8e1",
                          edgecolor="#c0392b", lw=0.8))
    fig.canvas.draw()
    bb = t.get_window_extent(renderer=fig.canvas.get_renderer())
    x0, y0 = inv.transform((bb.x0, bb.y0))
    x1, y1 = inv.transform((bb.x1, bb.y1))
    t.remove()
    return x1 - x0, y1 - y0


def _measure_text_plain(ax, text, fs):
    """测量文本（无边框）在数据坐标系中的实际渲染宽度 w"""
    fig = ax.figure
    inv = ax.transData.inverted()
    t = ax.text(0, 0, text, ha="left", va="center", fontsize=fs)
    fig.canvas.draw()
    bb = t.get_window_extent(renderer=fig.canvas.get_renderer())
    x0, _ = inv.transform((bb.x0, bb.y0))
    x1, _ = inv.transform((bb.x1, bb.y1))
    t.remove()
    return x1 - x0


def _name_fontsize(name_len):
    """信号名字号分档：保证 A4 打印清晰（最小 4.7pt）"""
    if name_len <= 12:
        return 7.0
    elif name_len <= 20:
        return 6.3
    elif name_len <= 28:
        return 5.5
    else:
        return 4.7


def _octagon_name_extent(ax, signals, OFF):
    """测所有信号名沿径向外伸的最大半径（数据单位）

    名字沿径向书写（长轴=径向），水平测量宽度即径向外伸长度。
    """
    verts, byte_edges = _octagon_geometry()
    max_r = OFF
    for sig in signals:
        segs = _split_byte_segments(sig)
        b, j0, j1 = segs[0]
        cmid = (j0 / 8 + (j1 + 1) / 8) / 2
        mx, my = _octagon_edge_point(verts, byte_edges, b, cmid)
        nx, ny = _octagon_outward(verts, byte_edges, b)
        fs = _name_fontsize(len(sig["name"]))
        name_len = _measure_text_plain(ax, sig["name"], fs)
        r = math.hypot(mx + nx * (OFF + name_len), my + ny * (OFF + name_len))
        max_r = max(max_r, r)
    return max_r


def plot_single_frame_octagon(ldf, frame_name, output_dir=OUTPUT_DIR, save_png=True, save_svg=True,
                              dpi=300, annotations=None, paper="landscape"):
    """绘制八边形布局单帧图（8条边=8字节，每条边8 bit，中心区帧信息）

    规则：
    - 字节按顺时针排列，byte0 顶边左→右，bit0 为该边最左端
    - 帧长 < 8 字节时，空余边以虚线 + "Reserved" 标记
    - 跨字节信号拆分为多段分别绘于各字节边，段间用弧线连接
    - 注释以引线+文本框绘制在八边形外侧（沿信号径向）
    paper: "landscape"(A4 横向, 默认) 或 "portrait"(A4 纵向)；仅改变纸张方向，
           八边形/注释列表的相对布局不变
    """
    from matplotlib.patches import FancyArrowPatch

    frame = ldf.get_unconditional_frame(frame_name)
    signals = get_signals(frame)
    total_bits = frame.length * 8
    n_bytes = frame.length

    verts, byte_edges = _octagon_geometry()

    # 颜色渐变（与线性布局一致）
    colors = ["#f0e68c", "#c5e1a5", "#81c784", "#4db6ac", "#26a69a",
              "#00897b", "#00695c", "#004d40", "#1a237e", "#0d47a1", "#ffcc80"]
    cmap = LinearSegmentedColormap.from_list("custom", colors, N=len(signals))

    # 收集注释
    ann_items = []
    if annotations:
        for sig in signals:
            text = (annotations.get(sig["name"]) or "").strip()
            if text:
                ann_items.append((sig, text))

    # 区块偏移参数（半径单位）
    OFF, W = 0.07, 0.16   # 信号块外偏移与宽度
    LABEL_R = OFF + W + 0.10   # Reserved 标签半径（外侧）
    BYTE_LABEL_R = 0.20        # 字节标签内偏移（向中心，避开径向信号名文字）

    # A4 纸张：横向(landscape, 默认) = 八边形左 + 注释列表右；
    #          纵向(portrait) = 八边形上 + 注释列表下（相对布局比例不变）
    if paper == "portrait":
        fig = plt.figure(figsize=(8.27, 11.69))          # A4 portrait
        ax = fig.add_axes([0.06, 0.42, 0.88, 0.54])      # 八边形区（上部）
        ax_ann = fig.add_axes([0.06, 0.02, 0.88, 0.36])  # 注释列表区（下部）
    else:
        fig = plt.figure(figsize=(11.69, 8.27))          # A4 landscape
        ax = fig.add_axes([0.01, 0.04, 0.46, 0.92])      # 八边形区
        ax_ann = fig.add_axes([0.50, 0.04, 0.49, 0.92])  # 注释列表区
    fig.patch.set_facecolor("white")
    ax.set_facecolor("#fafafa")
    ax.set_aspect("equal")
    ax_ann.set_xlim(0, 1)
    ax_ann.set_ylim(0, 1)
    ax_ann.set_xticks([])
    ax_ann.set_yticks([])
    for spine in ax_ann.spines.values():
        spine.set_visible(False)

    # 画幅自适应：信号名沿径向延伸，lim 由名称最大外伸决定
    # （字号固定 → 数据单位随变换缩放，与 lim 相互依赖 → 迭代收敛）
    lim = 1.6
    for _ in range(10):
        ax.set_xlim(-lim, lim)
        ax.set_ylim(-lim, lim)
        ext = _octagon_name_extent(ax, signals, OFF)
        new_lim = max(1.6, ext + 0.15)
        if abs(new_lim - lim) < 0.02:
            lim = new_lim
            break
        lim = new_lim
    ax.set_xlim(-lim, lim)
    ax.set_ylim(-lim, lim)

    # 八边形外轮廓（浅灰填充）
    poly = Polygon(verts, closed=True, facecolor="#f2f4f7", edgecolor="#555",
                   linewidth=1.2, zorder=1)
    ax.add_patch(poly)

    # 每条边的 bit 刻度（1~7 分隔线）与字节标签
    for b in range(8):
        nx, ny = _octagon_outward(verts, byte_edges, b)
        if b >= n_bytes:
            # Reserved 边：虚线 + 标签
            p0 = _octagon_edge_point(verts, byte_edges, b, 0.0)
            p1 = _octagon_edge_point(verts, byte_edges, b, 1.0)
            ax.plot([p0[0], p1[0]], [p0[1], p1[1]], color="#bbb", lw=1.0,
                    linestyle=(0, (4, 3)), zorder=2)
            mx, my = _octagon_edge_point(verts, byte_edges, b, 0.5)
            ax.text(mx + nx * LABEL_R, my + ny * LABEL_R, "Reserved",
                    ha="center", va="center", fontsize=7.5, color="#999",
                    fontstyle="italic", zorder=6)
            continue
        for j in range(1, 8):
            px, py = _octagon_edge_point(verts, byte_edges, b, j / 8)
            ax.plot([px + nx * (OFF - 0.02), px + nx * (OFF + W + 0.02)],
                    [py + ny * (OFF - 0.02), py + ny * (OFF + W + 0.02)],
                    color="#666", lw=0.6, zorder=3)
        # 字节标签的旋转角（沿边方向）
        ks, ke = byte_edges[b]
        dx = verts[ke][0] - verts[ks][0]
        dy = verts[ke][1] - verts[ks][1]
        ang = math.degrees(math.atan2(dy, dx))
        if ang > 90 or ang < -90:
            ang += 180
        # bit 标注：位于八边形边与信号块之间的间隙（径向偏移 = OFF/2），
        # 沿边方向书写，byte b 对应 bit b*8 ~ b*8+7
        BIT_LABEL_R = OFF * 0.5
        for j in range(8):
            bit_no = b * 8 + j
            px, py = _octagon_edge_point(verts, byte_edges, b, (j + 0.5) / 8)
            ax.text(px + nx * BIT_LABEL_R, py + ny * BIT_LABEL_R,
                    str(bit_no), ha="center", va="center",
                    rotation=ang, fontsize=5.0, color="#666", zorder=6)
        # 字节标签（放内侧靠近中心圆，避开径向信号名文字）
        mx, my = _octagon_edge_point(verts, byte_edges, b, 0.5)
        ax.text(mx - nx * BYTE_LABEL_R, my - ny * BYTE_LABEL_R,
                f"B{b} [{b * 8}-{b * 8 + 7}]", ha="center", va="center",
                rotation=ang, fontsize=7.5, fontweight="bold", color="#333",
                zorder=6, bbox=dict(boxstyle="round,pad=0.15", facecolor="white",
                                    edgecolor="#aaa", lw=0.5))

    # 绘制信号块（按字节分段）
    drawn = {}   # byte -> list[(t0, t1, color)] 用于跨字节连接
    for i, sig in enumerate(signals):
        color = cmap(i / max(len(signals) - 1, 1))
        segs = _split_byte_segments(sig)
        for idx, (b, j0, j1) in enumerate(segs):
            t0, t1 = j0 / 8, (j1 + 1) / 8
            quad = _octagon_block_quad(verts, byte_edges, b, t0, t1, OFF, W)
            ax.add_patch(Polygon(quad, closed=True, facecolor=color,
                                 edgecolor="#222", linewidth=0.6, alpha=0.92, zorder=4))
            drawn.setdefault(b, []).append((t0, t1, color))
            # 段内 bit 分隔
            for j in range(j0 + 1, j1 + 1):
                px, py = _octagon_edge_point(verts, byte_edges, b, j / 8)
                nx, ny = _octagon_outward(verts, byte_edges, b)
                ax.plot([px + nx * OFF, px + nx * (OFF + W)],
                        [py + ny * OFF, py + ny * (OFF + W)],
                        color="#222", lw=0.3, alpha=0.55, zorder=5)
            # 信号名（沿外法向[径向]书写，完整名称不缩写；
            # 锚定在信号块内边缘并只向外延伸，保证不越过八边形边、不遮挡内部）
            if idx == 0:
                cmid = (t0 + t1) / 2
                mx, my = _octagon_edge_point(verts, byte_edges, b, cmid)
                nx, ny = _octagon_outward(verts, byte_edges, b)
                # 文字长轴沿外法向（径向），规范化到 [-90,90] 保证不颠倒
                rad_ang = math.degrees(math.atan2(ny, nx))
                if rad_ang > 90 or rad_ang < -90:
                    rad_ang += 180
                # 外法向朝左(x<0)的边：文字由内端向左(外)延伸 → ha=right；其余 ha=left
                ha = "right" if nx < 0 else "left"
                # 字号按名称长度分档：径向文字沿垂直边书写（长轴=径向），
                # 不受沿边 seg_len 限制，始终显示完整名称、足够大
                fs = _name_fontsize(len(sig["name"]))
                # 锚点 = 信号块内边缘(径向 OFF)；rotation_mode="anchor" 使文字
                # 从锚点起只沿阅读方向延伸(ha=left 向前 / ha=right 向后=外法向)，
                # 保证信号名只位于八边形外侧、绝不越过八边形边
                ax.text(mx + nx * OFF, my + ny * OFF,
                        sig["name"], ha=ha, va="center", rotation=rad_ang,
                        rotation_mode="anchor",
                        fontsize=fs, fontweight="bold", color="#111", zorder=6)

    # 跨字节信号：段间弧线连接（在共享顶点外侧弯出）
    for i, sig in enumerate(signals):
        segs = _split_byte_segments(sig)
        if len(segs) < 2:
            continue
        color = cmap(i / max(len(signals) - 1, 1))
        for idx in range(len(segs) - 1):
            b1, j1a, j1b = segs[idx]
            b2, j2a, j2b = segs[idx + 1]
            p1 = _octagon_edge_point(verts, byte_edges, b1, (j1b + 1) / 8)
            p2 = _octagon_edge_point(verts, byte_edges, b2, j2a / 8)
            nx1, ny1 = _octagon_outward(verts, byte_edges, b1)
            nx2, ny2 = _octagon_outward(verts, byte_edges, b2)
            A = (p1[0] + nx1 * OFF, p1[1] + ny1 * OFF)
            B = (p2[0] + nx2 * OFF, p2[1] + ny2 * OFF)
            arc = FancyArrowPatch(A, B, arrowstyle="-", color=color, lw=1.8,
                                  connectionstyle="arc3,rad=-0.45",
                                  shrinkA=0, shrinkB=0, zorder=5)
            ax.add_patch(arc)

    # 中心区帧信息
    center = Circle((0, 0), 0.52, facecolor="white", edgecolor="#555",
                    linewidth=1.0, zorder=7)
    ax.add_patch(center)
    ax.text(0, 0.16, frame.name, ha="center", va="center", fontsize=11,
            fontweight="bold", color="#111", zorder=8)
    ax.text(0, -0.02, f"ID: {frame.frame_id}", ha="center", va="center",
            fontsize=8.5, color="#333", zorder=8)
    ax.text(0, -0.18, f"{n_bytes} B / {total_bits} bit", ha="center", va="center",
            fontsize=8.5, color="#555", zorder=8)

    # ===== 注释布局：右侧列表式（一条条顺序往下排列，自动分列） =====
    # 注释区独立于八边形区，远离信号名，天然不重叠
    if ann_items:
        sig_colors = {}
        for i, sig in enumerate(signals):
            sig_colors[sig["name"]] = cmap(i / max(len(signals) - 1, 1))
        ax_ann.text(0.02, 0.975, "信号注释", fontsize=13, fontweight="bold",
                    color="#111", transform=ax_ann.transAxes)
        entries, ann_fs, line_h = _layout_octagon_annotation_list(
            ax_ann, ann_items, sig_colors, fs_start=10.0, y_top=0.93)
        _draw_octagon_annotation_list(ax_ann, entries, ann_fs, line_h)

    ax.set_xticks([])
    ax.set_yticks([])
    for spine in ax.spines.values():
        spine.set_visible(False)

    fig.suptitle(f"Signal Mapping (Octagon) — {frame.name}  "
                 f"(Frame ID: {frame.frame_id}, Length: {frame.length} bytes)",
                 fontsize=15, fontweight="bold", y=0.985)

    saved = []
    if save_png:
        png_path = f"{output_dir}/{frame_name}_signal_mapping_octagon.png"
        plt.savefig(png_path, dpi=dpi, facecolor="white")
        saved.append(png_path)
    if save_svg:
        svg_path = f"{output_dir}/{frame_name}_signal_mapping_octagon.svg"
        plt.savefig(svg_path, format="svg", facecolor="white")
        saved.append(svg_path)
    plt.close()
    return saved


def plot_all_frames(ldf, frames_to_plot, output_dir=OUTPUT_DIR, save_png=True, save_svg=True, dpi=300):
    """绘制多帧总览图，返回生成的文件路径列表"""
    colors_list = ["#f0e68c", "#c5e1a5", "#81c784", "#4db6ac", "#26a69a", "#00897b",
                   "#00695c", "#004d40", "#1a237e", "#0d47a1", "#ffcc80", "#ffab91",
                   "#ce93d8", "#90caf9", "#a5d6a7"]

    fig, axes = plt.subplots(len(frames_to_plot), 1, figsize=(16, 3.5 * len(frames_to_plot)))
    if len(frames_to_plot) == 1:
        axes = [axes]

    for ax_idx, fname in enumerate(frames_to_plot):
        frame = ldf.get_unconditional_frame(fname)
        ax = axes[ax_idx]
        signals = get_signals(frame)
        total_bits = frame.length * 8

        # 网格
        for i in range(total_bits + 1):
            lw = 1.0 if i % 8 == 0 else 0.3
            color = "#222" if i % 8 == 0 else "#ddd"
            ax.axvline(i, color=color, linewidth=lw, ymin=0.2, ymax=0.8)

        # 信号块
        for i, sig in enumerate(signals):
            color = colors_list[i % len(colors_list)]
            rect = Rectangle((sig["start"], 0.3), sig["size"], 0.4,
                             facecolor=color, edgecolor="black", linewidth=0.7, alpha=0.85)
            ax.add_patch(rect)

            # 信号完整标识符（不简写）
            draw_signal_label(ax, sig, 0.5)

        # 字节号
        for i in range(frame.length):
            ax.text(i * 8 + 4, 0.92, str(i), ha="center", va="center", fontsize=9, fontweight="bold")

        # bit 号
        for i in range(0, total_bits, 8):
            ax.text(i, 0.12, str(i), ha="center", va="top", fontsize=7)
        ax.text(total_bits, 0.12, str(total_bits - 1), ha="center", va="top", fontsize=7)

        ax.set_title(f"{fname}  (ID={frame.frame_id}, {frame.length} bytes) — Publisher: {frame.publisher.name}",
                     fontsize=11, fontweight="bold", loc="left")
        ax.set_xlim(-0.5, total_bits + 0.5)
        ax.set_ylim(0, 1.05)
        ax.set_yticks([])
        ax.set_xticks([])
        for spine in ax.spines.values():
            spine.set_visible(False)

    plt.suptitle("LIN Signal Mapping Visualization (from LDF)", fontsize=14, fontweight="bold", y=1.01)
    plt.tight_layout()

    saved = []
    if save_png:
        png_path = f"{output_dir}/LIN_Signal_Mappings_All_Frames.png"
        plt.savefig(png_path, dpi=dpi, bbox_inches="tight", facecolor="white")
        saved.append(png_path)
    if save_svg:
        svg_path = f"{output_dir}/LIN_Signal_Mappings_All_Frames.svg"
        plt.savefig(svg_path, format="svg", bbox_inches="tight", facecolor="white")
        saved.append(svg_path)
    plt.close()
    return saved


def export_excel(ldf, frames_to_plot, output_dir=OUTPUT_DIR):
    """生成 Excel 汇总（仅包含指定帧），返回文件路径

    布局目标：适配 A4 页面（纸张/边距/缩放设置），拷贝到 Word 后不超宽
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    excel_path = f"{output_dir}/LIN_Signal_Mapping.xlsx"
    wb = Workbook()

    # 样式
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    body_font = Font(name="Arial", size=10)
    thin_border = Border(*[Side(style="thin", color="B0B0B0")] * 4)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # A4 纵向可打印宽约 82 个字符单位，Word 默认正文宽度相当；留余量取 80
    A4_WIDTH_UNITS = 80
    MIN_COL_WIDTH = 6

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

    def style_body(ws, nrows, ncols):
        for r in range(2, nrows + 1):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = body_font
                cell.border = thin_border
                cell.alignment = left if c in (1, 2, 3) else center

    def auto_width(ws, ncols):
        """按内容计算列宽：总宽固定为 A4_WIDTH_UNITS，按内容长度比例分配
        （短列保底 MIN_COL_WIDTH，长文本靠 wrap_text 换行显示完整内容）"""
        raw = []
        for c in range(1, ncols + 1):
            max_len = 0
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    # 中文/全角字符按 2 个单位估算
                    length = sum(2 if ord(ch) > 0xFF else 1 for ch in str(v))
                    max_len = max(max_len, length)
            raw.append(max(max_len + 3, MIN_COL_WIDTH))

        # 比例分配：每列先保底 MIN_COL_WIDTH，剩余宽度按 (raw - MIN) 占比分配
        min_total = MIN_COL_WIDTH * ncols
        extra_budget = max(A4_WIDTH_UNITS - min_total, 0)
        extra_weights = [max(r - MIN_COL_WIDTH, 0) for r in raw]
        weight_sum = sum(extra_weights)

        widths = []
        for i, w in enumerate(raw):
            if weight_sum and extra_budget:
                extra = extra_weights[i] / weight_sum * extra_budget
            else:
                extra = 0
            widths.append(MIN_COL_WIDTH + extra)

        # 由于四舍五入可能差 1-2 个单位，补偿到内容最多的列
        deficit = A4_WIDTH_UNITS - sum(widths)
        if deficit != 0:
            top = max(range(ncols), key=lambda i: extra_weights[i])
            widths[top] = max(widths[top] + deficit, MIN_COL_WIDTH)

        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = round(w, 1)

    def setup_a4(ws):
        """A4 页面设置：纸张、边距、缩放适配一页宽、打印重复表头"""
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6,
                                      header=0.3, footer=0.3)
        ws.print_title_rows = "1:1"
        ws.print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # ---- Sheet 1: Frames 概览 ----
    ws_frames = wb.active
    ws_frames.title = "Frames"
    frames_headers = ["Frame Name", "Frame ID", "Length (bytes)", "Length (bits)", "Publisher", "Signal Count"]
    ws_frames.append(frames_headers)
    for frame in ldf.get_unconditional_frames():
        if frame.name in frames_to_plot:
            ws_frames.append([frame.name, frame.frame_id, frame.length, frame.length * 8,
                              frame.publisher.name, len(frame.signal_map)])
    style_header(ws_frames, len(frames_headers))
    style_body(ws_frames, ws_frames.max_row, len(frames_headers))
    auto_width(ws_frames, len(frames_headers))
    setup_a4(ws_frames)

    # ---- Sheet 2: Signals 明细 ----
    ws_sig = wb.create_sheet("Signals")
    sig_headers = ["Frame Name", "Frame ID", "Signal Name", "Start Bit", "Size (bits)", "End Bit",
                   "Init Value", "Encoding Type", "Publisher", "Subscribers"]
    ws_sig.append(sig_headers)
    for frame in ldf.get_unconditional_frames():
        if frame.name not in frames_to_plot:
            continue
        for sig in get_signals(frame):
            lin_sig = next(s for s in frame.signal_map if s[0] == sig["start"])[1]
            ws_sig.append([
                frame.name,
                frame.frame_id,
                sig["name"],
                sig["start"],
                sig["size"],
                sig["end"],
                lin_sig.init_value,
                lin_sig.encoding_type.name if lin_sig.encoding_type else "",
                lin_sig.publisher.name if lin_sig.publisher else "",
                ", ".join(s.name for s in lin_sig.subscribers) if lin_sig.subscribers else "",
            ])
    style_header(ws_sig, len(sig_headers))
    style_body(ws_sig, ws_sig.max_row, len(sig_headers))
    auto_width(ws_sig, len(sig_headers))
    setup_a4(ws_sig)

    wb.save(excel_path)
    return excel_path


def main():
    """CLI 入口：解析默认 LDF 并生成全部默认帧"""
    ldf = ldfparser.parse_ldf(path=LDF_PATH, encoding="latin-1")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print(f"Baudrate: {ldf.get_baudrate()} bps")
    print("Unconditional Frames:")
    for frame in ldf.get_unconditional_frames():
        print(f"  {frame.name} (ID={frame.frame_id}, {frame.length} bytes, Publisher={frame.publisher.name})")

    # 1. 单帧高清图
    for fname in DEFAULT_FRAMES:
        saved = plot_single_frame(ldf, fname, OUTPUT_DIR)
        for s in saved:
            print(f"Saved: {s}")

    # 1b. 单帧八边形布局图（8边=8字节）
    for fname in DEFAULT_FRAMES:
        saved = plot_single_frame_octagon(ldf, fname, OUTPUT_DIR)
        for s in saved:
            print(f"Saved: {s}")

    # 2. 多帧总览图
    saved = plot_all_frames(ldf, DEFAULT_FRAMES, OUTPUT_DIR)
    for s in saved:
        print(f"Saved: {s}")

    # 3. Excel
    try:
        path = export_excel(ldf, DEFAULT_FRAMES, OUTPUT_DIR)
        print(f"Saved: {path}")
    except PermissionError:
        print("错误: LIN_Signal_Mapping.xlsx 被占用（可能在 Excel 中打开），请关闭后重试")

    print("\nDone!")


if __name__ == "__main__":
    main()
